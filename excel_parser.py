"""Aluminyum dograma siparis excelini (Genel sayfasi) parse eden ETL modulu.

Excel sabit satir numaralarina sahip degildir: her siparis dosyasinda blok
satir sayilari degisebilir. Bu yuzden bloklar B kolonundaki sabit Ingilizce
etiketlerle (Profiles/Hardware/Articles/Gaskets/Glass/Sum) bulunur, footer
ozet degerleri de satir numarasina degil etiket metnine gore aranir.

Bu modul Flask/DB'den bagimsizdir: parse_order_excel() sadece dict/list
dondurur, cagiran taraf (route) bunlardan Order/OrderItem olusturur.
"""

from datetime import datetime

import openpyxl

from models import ItemCategory

DEFAULT_SHEET_NAME = 'Genel'

BLOCK_TAG_TO_CATEGORY = {
    'Profiles': ItemCategory.PROFILE,
    'Hardware': ItemCategory.HARDWARE,
    'Articles': ItemCategory.CONSUMABLE,
    'Gaskets': ItemCategory.GASKET,
    'Glass': ItemCategory.GLASS,
}
SUM_TAG = 'Sum'

# Ortak kolon haritasi: Profiller / Aksesuar / Sarf malzemeler / Fitiller
COMMON_COLUMN_MAP = {
    'quantity': 'C',
    'length': 'H',
    'total_quantity': 'I',
    'unit': 'M',
    'stock_code': 'N',
    'stock_name': 'P',
    'color': 'Q',
    'unit_price': 'T',
    'total_price': 'W',
}

# Cam bloguna ozel kolon haritasi (stok kodu yok, olcu alanlari var)
GLASS_COLUMN_MAP = {
    'quantity': 'C',
    'width_mm': 'H',
    'height_mm': 'I',
    'area_m2': 'N',
    'stock_name': 'P',
    'thickness_mm': 'Q',
    'unit_price': 'T',
    'total_price': 'W',
}

# Footer ozet: (etiket alt-metni, aranacak kolon, deger kolonu) -> Order alani
FOOTER_LABEL_MAP = [
    ('KDV HARİÇ NAKİT MALZEME TUTARI', 'Q', 'W', 'kdv_haric_malzeme_tutari'),
    ('ALÜMİNYUM KG', 'C', 'I', 'aluminyum_kg'),
    ('İMALAT', 'S', 'W', 'iscilik_tutari'),
    ('ALÜMİNYUM MALZEME', 'C', 'I', 'aluminyum_malzeme_tutari'),
    ('AKSESUAR', 'C', 'I', 'aksesuar_tutari'),
    ('KAR', 'U', 'V', 'kar_orani'),
    ('SARF', 'C', 'I', 'sarf_tutari'),
    ('KDV HARİÇ NAKİT ATÖLYE TESLİM TUTARI', 'Q', 'W', 'kdv_haric_atolye_teslim_tutari'),
    ('FİTİL', 'C', 'I', 'fitil_tutari'),
    ('KDV HARİÇ CAMSIZ MALİYET', 'C', 'I', 'kdv_haric_camsiz_maliyet'),
]


class ExcelParseError(Exception):
    """Excel dosyasi beklenen siparis formatinda okunamadiginda firlatilir."""


def parse_order_excel(filepath, sheet_name=DEFAULT_SHEET_NAME):
    """Siparis excelini ayristirir.

    Returns:
        tuple[dict, list[dict]]: (order_fields, order_items)
    Raises:
        ExcelParseError: dosya okunamiyorsa veya beklenen yapi bulunamazsa.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f'Dosya okunamadi: {exc}') from exc

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    tag_rows = _scan_column_b(ws)
    if not tag_rows:
        raise ExcelParseError('Beklenen siparis formati bulunamadi (blok etiketleri yok)')

    order_fields = _read_header_fields(ws)
    order_fields.update(_read_footer_fields(ws, tag_rows))

    order_items = []
    for category_tag, category in BLOCK_TAG_TO_CATEGORY.items():
        data_rows = _block_data_rows(tag_rows, category_tag)
        column_map = GLASS_COLUMN_MAP if category_tag == 'Glass' else COMMON_COLUMN_MAP
        for row in data_rows:
            item = _read_item_row(ws, row, column_map, category)
            if item is not None:
                order_items.append(item)

    order_fields['total_amount'] = order_fields.get('kdv_haric_atolye_teslim_tutari') or 0.0

    return order_fields, order_items


def _scan_column_b(ws):
    """B kolonunda gecen (satir, etiket) ciftlerini sirayla dondurur."""
    recognized = set(BLOCK_TAG_TO_CATEGORY.keys()) | {SUM_TAG}
    tag_rows = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value  # B kolonu
        if isinstance(value, str) and value.strip() in recognized:
            tag_rows.append((row, value.strip()))
    return tag_rows


def _block_data_rows(tag_rows, category_tag):
    """Belirli bir kategori etiketinin ard arda gectigi satir numaralarini dondurur."""
    return [row for row, tag in tag_rows if tag == category_tag]


def _read_item_row(ws, row, column_map, category):
    """Tek bir veri satirindan OrderItem dict'i uretir. Bos satirlarda None doner."""
    stock_name = _cell(ws, row, column_map.get('stock_name'))
    stock_code = _cell(ws, row, column_map.get('stock_code')) if 'stock_code' in column_map else None
    if not stock_name and not stock_code:
        return None

    item = {'category': category, 'status': 'Bekliyor'}

    item['quantity'] = _to_float(_cell(ws, row, column_map.get('quantity')), default=0.0)
    item['total_quantity'] = _to_float(_cell(ws, row, column_map.get('total_quantity')), default=0.0)
    item['unit'] = _clean_str(_cell(ws, row, column_map.get('unit')))
    item['stock_code'] = _clean_str(stock_code)
    item['stock_name'] = _clean_str(stock_name)
    item['unit_price'] = _to_float(_cell(ws, row, column_map.get('unit_price')), default=0.0)
    item['total_price'] = _to_float(_cell(ws, row, column_map.get('total_price')), default=0.0)

    if 'length' in column_map:
        raw_length = _cell(ws, row, column_map['length'])
        item['length'] = str(raw_length).strip() if raw_length not in (None, '') else None

    if 'color' in column_map:
        item['color'] = _clean_str(_cell(ws, row, column_map.get('color')))

    if 'width_mm' in column_map:
        item['width_mm'] = _to_float(_cell(ws, row, column_map.get('width_mm')), default=None)
    if 'height_mm' in column_map:
        item['height_mm'] = _to_float(_cell(ws, row, column_map.get('height_mm')), default=None)
    if 'thickness_mm' in column_map:
        item['thickness_mm'] = _to_float(_cell(ws, row, column_map.get('thickness_mm')), default=None)
    if 'area_m2' in column_map:
        item['area_m2'] = _to_float(_cell(ws, row, column_map.get('area_m2')), default=None)

    return item


def _read_header_fields(ws):
    """Ust bilgi (Tarih/Proje Adi/Sorumlu Kisi) alanlarini okur."""
    project_name = _clean_str(ws['W2'].value)
    if not project_name:
        raise ExcelParseError('Proje adi okunamadi (W2 hucresi bos)')

    order_date = None
    raw_date = ws['W1'].value
    if isinstance(raw_date, datetime):
        order_date = raw_date.date()
    elif isinstance(raw_date, str) and raw_date.strip():
        try:
            order_date = datetime.strptime(raw_date.strip(), '%d.%m.%Y').date()
        except ValueError:
            order_date = None

    return {
        'order_date': order_date,
        'project_name': project_name,
        'responsible_person': _clean_str(ws['W3'].value),
    }


def _read_footer_fields(ws, tag_rows):
    """Footer ozet degerlerini (KDV haric tutarlar, aluminyum kg vb.) etiket bazli okur."""
    sum_rows = [row for row, tag in tag_rows if tag == SUM_TAG]
    start_row = (max(sum_rows) if sum_rows else 1) + 1

    fields = {}
    for label, label_col, value_col, field_name in FOOTER_LABEL_MAP:
        cell = _find_label_cell(ws, label, label_col, start_row)
        if cell is None:
            continue
        value = ws[f'{value_col}{cell.row}'].value
        fields[field_name] = _to_float(value, default=None)
    return fields


def _find_label_cell(ws, label, column_letter, start_row):
    """Belirli bir kolonda, start_row'dan itibaren verilen etiketi iceren hucreyi bulur."""
    label_upper = label.upper()
    for row in range(start_row, ws.max_row + 1):
        value = ws[f'{column_letter}{row}'].value
        if isinstance(value, str) and label_upper in value.strip().upper():
            return ws[f'{column_letter}{row}']
    return None


def _cell(ws, row, column_letter):
    if column_letter is None:
        return None
    return ws[f'{column_letter}{row}'].value


def _clean_str(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_float(value, default=0.0):
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


if __name__ == '__main__':
    fields, items = parse_order_excel(r'c:\Software\siparistest.XLSX')
    print('ORDER FIELDS:')
    for k, v in fields.items():
        print(f'  {k}: {v!r}')
    print(f'\nITEM COUNT: {len(items)}')
    by_category = {}
    for it in items:
        by_category.setdefault(it['category'], []).append(it)
    for cat, cat_items in by_category.items():
        print(f'\n[{cat}] {len(cat_items)} kalem')
        for it in cat_items:
            print(f"  - {it.get('stock_code')!r:20} {it.get('stock_name')!r:50} adet={it.get('quantity')} toplam={it.get('total_price')}")
